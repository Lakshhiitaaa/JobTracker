"""Formatted Excel output."""
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import config

COLUMNS = ["Company", "Job Title", "Category", "Experience", "Location",
           "Date Found", "Apply Link", "Careers Page", "New Today"]
WIDTHS = [26, 44, 16, 20, 24, 12, 46, 40, 11]

HDR_FILL = PatternFill("solid", fgColor="1F3864")
NEW_FILL = PatternFill("solid", fgColor="DFF5E1")
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def build(rows, errors=None, out_dir=None):
    out_dir = out_dir or config.REPORTS_DIR
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"jobs_{date.today().isoformat()}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"
    ws.append(COLUMNS)
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    # new jobs first
    rows = sorted(rows, key=lambda r: (not r.get("is_new"), r.get("company", "")))
    for r in rows:
        ws.append([r.get("company", ""), r.get("title", ""), r.get("category", ""),
                   r.get("experience", ""), r.get("location", ""),
                   r.get("date_found", date.today().isoformat()),
                   r.get("url", ""), r.get("careers_page", ""),
                   "Yes" if r.get("is_new") else "No"])
        row_i = ws.max_row
        if r.get("url"):
            cell = ws.cell(row=row_i, column=7)
            cell.hyperlink = r["url"]
            cell.font = Font(color="0563C1", underline="single")
        for col in range(1, len(COLUMNS) + 1):
            ws.cell(row=row_i, column=col).border = THIN
            if r.get("is_new"):
                ws.cell(row=row_i, column=col).fill = NEW_FILL
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(ws.max_row, 1)}"

    if errors:
        we = wb.create_sheet("Errors")
        we.append(["Company", "Careers URL", "Error"])
        for i, w in enumerate([26, 50, 70], 1):
            we.column_dimensions[get_column_letter(i)].width = w
            we.cell(row=1, column=i).font = Font(bold=True, color="FFFFFF")
            we.cell(row=1, column=i).fill = HDR_FILL
        for e in errors:
            we.append([e.get("company", ""), e.get("url", ""), str(e.get("error", ""))[:300]])

    wb.save(path)
    return path
