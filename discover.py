"""Find and verify the real ATS board URL for a company.

  python discover.py https://zerodha.com/careers/     one URL
  python discover.py --name "Hasura"                  guess tokens from a name
  python discover.py --audit                          check every row in companies.xlsx

--audit writes reports/company_audit.xlsx telling you, per company, whether it
resolves to a working ATS API and what URL to paste into companies.xlsx.
"""
import argparse, re, sys, os
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
import config, ats

HEADERS = {"User-Agent": config.USER_AGENT}

CANONICAL = {
    "greenhouse":      "https://boards.greenhouse.io/{t}",
    "lever":           "https://jobs.lever.co/{t}",
    "ashby":           "https://jobs.ashbyhq.com/{t}",
    "smartrecruiters": "https://careers.smartrecruiters.com/{t}",
    "workable":        "https://apply.workable.com/{t}",
    "recruitee":       "https://{t}.recruitee.com",
}


def verify(ats_name, token):
    """Call the ATS API. Returns (ok: bool, job_count: int, note: str)."""
    if ats_name not in ats.CLIENTS:
        return False, 0, "detected but no JSON client (falls back to HTML)"
    try:
        jobs = ats.fetch(ats_name, token)
        if not jobs:
            return False, 0, "API responded but returned 0 jobs"
        return True, len(jobs), "ok"
    except Exception as e:
        return False, 0, f"{type(e).__name__}: {str(e)[:80]}"


def token_guesses(name_or_url):
    """Plausible board slugs from a company name or domain."""
    s = name_or_url.lower()
    m = re.search(r"https?://(?:www\.)?([a-z0-9-]+)\.", s)
    if m:
        s = m.group(1)
    s = re.sub(r"\b(technologies|technology|labs|inc|pvt|ltd|limited|india)\b", "", s)
    base = re.sub(r"[^a-z0-9]", "", s)
    hyph = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    out = [base, hyph, base + "inc", base + "india", base + "technologies"]
    seen, uniq = set(), []
    for t in out:
        if t and t not in seen:
            seen.add(t); uniq.append(t)
    return uniq


def brute_force(name_or_url, quiet=False):
    """Try every ATS x every token guess. Returns list of working (ats, token, count)."""
    hits = []
    guesses = token_guesses(name_or_url)
    for ats_name in CANONICAL:
        for tok in guesses:
            ok, n, _ = verify(ats_name, tok)
            if ok:
                hits.append((ats_name, tok, n))
                if not quiet:
                    print(f"  FOUND  {ats_name:<16} token={tok:<20} {n} jobs")
    return hits


def inspect(url, quiet=False):
    """Full check of one careers URL. Returns dict."""
    result = {"url": url, "ats": None, "token": None, "verified": False,
              "jobs": 0, "suggested": "", "note": ""}
    html = ""
    try:
        r = requests.get(url, headers=HEADERS, timeout=config.REQUEST_TIMEOUT)
        html = r.text
    except Exception as e:
        result["note"] = f"page fetch failed: {type(e).__name__}"

    name, token = ats.detect(url, html)
    if name:
        result["ats"], result["token"] = name, token
        ok, n, note = verify(name, token)
        result["verified"], result["jobs"], result["note"] = ok, n, note
        if ok:
            result["suggested"] = CANONICAL.get(name, url).format(t=token)
            if not quiet:
                print(f"  {name} / {token} -> {n} jobs")
            return result
        if not quiet:
            print(f"  detected {name}/{token} but {note}; trying other tokens...")

    hits = brute_force(url, quiet)
    if hits:
        a, t, n = max(hits, key=lambda x: x[2])
        result.update({"ats": a, "token": t, "verified": True, "jobs": n,
                       "suggested": CANONICAL[a].format(t=t), "note": "found by guessing"})
    elif not result["note"]:
        result["note"] = "no ATS found - will use HTML/Playwright scraping"
    return result


def audit():
    import pandas as pd
    from openpyxl.styles import Font, PatternFill
    df = pd.read_excel(config.COMPANIES_FILE)
    df.columns = [str(c).strip() for c in df.columns]
    rows = df.to_dict("records")
    print(f"Auditing {len(rows)} companies...\n")

    out = []
    with ThreadPoolExecutor(max_workers=config.MAX_WORKERS) as ex:
        futs = {ex.submit(inspect, str(r["Careers URL"]).strip(), True): r for r in rows}
        for i, f in enumerate(as_completed(futs), 1):
            r = futs[f]
            name = str(r["Company Name"])
            try:
                res = f.result()
            except Exception as e:
                res = {"ats": None, "token": None, "verified": False, "jobs": 0,
                       "suggested": "", "note": str(e)[:80]}
            flag = "OK  " if res["verified"] else "FIX "
            print(f"  [{i}/{len(rows)}] {flag} {name:<28} {str(res['ats'] or '-'):<16} {res['jobs']:>3} jobs")
            out.append({
                "Company Name": name,
                "Current URL": r["Careers URL"],
                "Detected ATS": res["ats"] or "",
                "Token": res["token"] or "",
                "Verified": "Yes" if res["verified"] else "No",
                "Jobs on Board": res["jobs"],
                "Suggested URL": res["suggested"],
                "Note": res["note"],
            })

    os.makedirs(config.REPORTS_DIR, exist_ok=True)
    path = os.path.join(config.REPORTS_DIR, "company_audit.xlsx")
    adf = pd.DataFrame(out).sort_values(["Verified", "Company Name"])
    adf.to_excel(path, index=False)

    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for i, w in enumerate([26, 46, 16, 20, 10, 13, 46, 50], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.cell(row=1, column=i).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=1, column=i).fill = PatternFill("solid", fgColor="1F3864")
    ws.freeze_panes = "A2"
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=5).value == "No":
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FDE9E9")
    wb.save(path)

    good = sum(1 for r in out if r["Verified"] == "Yes")
    print(f"\n{good}/{len(out)} on a verified ATS. Audit: {path}")
    print("Red rows need attention. Copy 'Suggested URL' into companies.xlsx where present.")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("url", nargs="?", help="a careers page URL")
    ap.add_argument("--name", help="company name, guess tokens without a URL")
    ap.add_argument("--audit", action="store_true", help="check all of companies.xlsx")
    a = ap.parse_args()

    if a.audit:
        audit()
    elif a.name:
        print(f"Guessing tokens for '{a.name}': {', '.join(token_guesses(a.name))}")
        if not brute_force(a.name):
            print("  nothing found on any supported ATS")
    elif a.url:
        r = inspect(a.url)
        print()
        if r["verified"]:
            print(f"  ATS       : {r['ats']}")
            print(f"  Token     : {r['token']}")
            print(f"  Jobs      : {r['jobs']}")
            print(f"  USE THIS  : {r['suggested']}")
        else:
            print(f"  No verified ATS. {r['note']}")
            print("  Leave the marketing careers URL in place; HTML scraping will try.")
    else:
        ap.print_help(); sys.exit(1)
